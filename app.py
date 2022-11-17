from flask import Flask, render_template, jsonify, request, Response
from flask_restful import Api, Resource, reqparse
from flask_sqlalchemy import SQLAlchemy
from flask_marshmallow import Marshmallow
import urllib.request, urllib.parse, json
from abe import main
from time import *
import numpy as np
import socket
from datetime import datetime
import csv
import config
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import os
from pytz import timezone
from db_models import *

app = Flask(__name__, static_folder='static')
api = Api(app)

app.config["SECRET_KEY"] = "abefmscanner"
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///abe/get_status.db'

db.app = app
db.init_app(app)

urlbase= config.urlbase


########CSV MANAGE ############################################################################################################################

CSVfilePath = config.CSVfilePath
tunnelName = config.tunnelName

def createCSVfile(path, now , tunnelName):
    wb = Workbook()
    ws = wb.active
    fileName = now + "-"+ tunnelName
    ws.title = "Tunnel Data"
    return wb, ws, fileName

def getTime():
    turkey = timezone('Europe/Istanbul')
    now = datetime.now(turkey)
    fileNow =now.strftime("%Y.%m.%d-%H.%M.%S")
    return fileNow

def deleteCSVfile(path):
    for file in os.listdir(path):  
        if (file.endswith('.xlsx')):
            try:
                os.remove(path+file)
            except Exception as e:
                print(e)
        else:
            pass

class manageExcel():
    def __init__(self, wb,ws, fileName):
        self.wb = wb
        self.ws = ws
        self.fileName = fileName

    def appendHeader(self):
        wb = self.wb
        ws = self.ws
        fileName = self.fileName
        ws.append(['Date', 'Frequency', 'Validness', 'Signal Power', 'SNR', 'Temperature', 'Humidity'])
        wb.save(fileName +'.xlsx')

    def appendData(self, CSVdata):
        wb = self.wb
        ws = self.ws
        fileName = self.fileName
        for data in CSVdata:
            try:
                ws.append([data.date, data.freq, data.resp1, data.rssi, data.snr, data.temp, data.hum])
            except Exception as e:
                print(e)
        wb.save(fileName +'.xlsx')


###########################################################################################################################################

global CurrentChannel
CurrentChannel = 0
Frequency_Scan = main.SerialCommunication()

def getFrequencies():
    url= urlbase + "/frequencyData"
    response = urllib.request.urlopen(url)
    data =response.read()
    dict =json.loads(data)
    return dict

# def postFrequencyData():
#     data = { 'test1': 10, 'test2': 20 }
#     data = urllib.parse.urlencode(data).encode()
#     req =  urllib.request.Request( urlbase + "/frequencyData", data=data) # this will make the method "POST"
#     resp = urllib.request.urlopen(req)
#     print(resp.read())

        
def updateFrequencyList():
    try:
        frequencyData = getFrequencies() 
        db.session.query(status).delete()
        db.session.commit()

        for frequency in frequencyData:
            if frequency['status'] == 1:
                freq = frequency['freq']
                kanal = frequency['name']
                resp1 = 0
                rssi = 0
                snr = 0
                stat = " "
                new_frequency = status(freq, kanal, resp1, rssi, snr, stat)
                db.session.add(new_frequency)
                db.session.commit()
            else:
                pass

    except Exception as e:
        print(e)


# SocketFlag = False ###https://stackoverflow.com/questions/48024720/python-how-to-check-if-socket-is-still-connected

updateFrequencyList()

@app.route('/audio', methods =['GET', 'POST'])
def audio():

     ##   if SocketFlag= is_socket_closed() #bağlı cihaz yoksa true cönüp alttaki fonksiyonu çalıştırcak. else meşgul uyarısı versin.
    try:
        # socketflag = true ##bu gereksiz olabilir bi dnemek lazım. 
        sleep(0.5)
        s =socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        s.connect((socket.gethostname(),1235))
        print(id(s))

        def sound():
            first_run = True
            print("recording...")
            while True:
                try:
                    if first_run :
                        data =s.recv(4096)
                        first_run = False
                    else:
                        data =s.recv(4096)
                        data_array = np.frombuffer(data, dtype='int16')
                        channel0 = data_array[0::2]
                        data = channel0.tostring()   
                    yield(data)      
                          
                except Exception as e:
                    first_run = True
                    print(e)
                    

    except Exception as e:
        print(e)

    return Response(sound())

@app.route("/", methods=['GET', 'POST'])
def MainPage():
    # if False:  # Buraya error şartı gelecek, true ise error sayfasına gidecek
    #     return render_template("errorpage.html")
    global CurrentChannel
    Frequency_Scan.Module_1_One_Frequency(CurrentChannel)
    items = status.query.all()
    mediums = medium.query.all()
    hists = records.query.all()
    text = [item.freq for item in items]
    if request.method == 'POST':
        for task in request.form:
            if task == 'next':
                if (CurrentChannel+1) < len(items):
                    CurrentChannel = CurrentChannel + 1                               
                else:
                    CurrentChannel = 0
                
                Frequency_Scan.Module_1_One_Frequency(CurrentChannel)

            elif task == 'prev':
                if  CurrentChannel != 0:
                    CurrentChannel = CurrentChannel - 1
                    print(CurrentChannel)                
                else:
                    CurrentChannel = len(items) - 1
                
                Frequency_Scan.Module_1_One_Frequency(CurrentChannel)

            elif task == 'Kanal Listesini Güncelle':
                updateFrequencyList()
            else:
                ClickedChannel = int(request.form["kanalFrekansı"])
                for t in text:
                    if ClickedChannel == t:
                        CurrentChannel = text.index(t)
                    else:
                        pass
                Frequency_Scan.Module_1_One_Frequency(CurrentChannel)

    return render_template('mainpage.html', items = items, CurrentChannel = CurrentChannel, mediums = mediums, text = text, hists = hists) #socketflagi buraya ekle

@app.route('/records', methods=['GET', 'POST'])
def showRecords():
    deleteCSVfile(CSVfilePath)
    fileNow = getTime()
    wb, ws, fileName =createCSVfile(CSVfilePath, fileNow , tunnelName)
    CSV = manageExcel(wb, ws, fileName)
    CSV.appendHeader()
    CSVdata = data.query.all()  
    CSV.appendData(CSVdata)
    hists = records.query.all()
    if request.method == 'POST':
        for task in request.form:
            if task == 'Apply':
                deleteCSVfile(CSVfilePath)
                fileNow = getTime()
                wb, ws, fileName =createCSVfile(CSVfilePath, fileNow , tunnelName)
                CSV = manageExcel(wb, ws, fileName)
                CSV.appendHeader()
                
                filteredCSVdata =[]
                newHists =[]
                fromDatestr = request.form["from_Date"]
                fromDate = datetime.strptime( fromDatestr , '%Y-%m-%d')
                toDatestr = request.form["to_Date"]
                toDate = datetime.strptime( toDatestr + " " + "23:59:59", '%Y-%m-%d %H:%M:%S')

                CSVdatafilter = data.query.all() 

                for info in CSVdatafilter:
                    date = datetime.strptime(info.date, '%Y-%m-%d %H:%M:%S.%f')
                    if date >= fromDate and date <= toDate:
                        filteredCSVdata.append(info)

                CSV.appendData(filteredCSVdata)
                for hist in hists:
                    date = datetime.strptime(hist.name, '%Y-%m-%d %H:%M:%S')
                    if date >= fromDate and date <= toDate:
                        newHists.append(hist)
                return  render_template('recordings.html', hists = newHists,  CSVfileName =CSVfilePath + fileName + ".xlsx")   
            
            elif task == 'submit':
                return render_template('recordings.html')

    return render_template('recordings.html', hists = hists, CSVfileName =CSVfilePath + fileName + ".xlsx")

#############################################################API END POINTS##########################################################################################

#class for manipulating only one channel entry and its parameters#######################
class Frequency(Resource):
#gets data of chosen frequency
    def get(self, frequency):
        parameters = status.query.get(frequency)
        return parameter_schema.jsonify(parameters)

#adds new frequency to database 
    def post(self,frequency):    

        freq = request.json['freq']
        kanal = request.json['kanal']
        resp1 = 0
        rssi = 0
        snr = 0
        stat = " "

        new_frequency = status(freq, kanal, resp1, rssi, snr, stat)

        db.session.add(new_frequency)
        db.session.commit()

        return parameter_schema.jsonify(new_frequency)
#updates all data of chosen frequency
    def put(self, frequency):
        parameters = status.query.get(frequency)

        freq = request.json['freq']
        kanal = request.json['kanal']
        # resp1 = request.json['resp1']
        # rssi = request.json['rssi']
        # snr = request.json['snr']
        # stat = request.json['stat']

        parameters.freq = freq
        parameters.kanal = kanal
        # parameters.resp1 = resp1
        # parameters.rssi = rssi
        # parameters.snr = snr
        # parameters.stat = stat

        db.session.commit()

        return parameter_schema.jsonify(parameters)

#deletes chosen frequency from database
    def delete(self, frequency):
        parameters = status.query.get(frequency)
        db.session.delete(parameters)
        db.session.commit()
        return parameter_schema.jsonify(parameters)

# class for manipulating all channel datas and their parameters###################################
class allFrequencies(Resource):
    def get(self):
        all_Frequencies = status.query.all()
        result = parameters_schema.dump(all_Frequencies)
        return jsonify(result)
    
    def post(self):
        return {}

    def put(self):
        return{}
    
    def delete(self):
        return {}

#class for manipulating temperature and humidity data that taken from device regularly##############
class temperatureAndHumidity(Resource):
    def get(self):
        parameters = medium.query.all()
        result = medium_schema.dump(parameters)
        return jsonify(result)
    
    def post(self):
        date = request.json['date']
        temp = request.json['temp']
        hum = request.json['hum']

        new_medium = status(date,temp,hum)

        db.session.add(new_medium)
        db.session.commit()

        return medium_schema.jsonify(new_medium)


    def put(self):
        return{}

    def delete(self):
        return {}

class isThereSound(Resource):
    def get(self):
        return{}

class isThereConnection(Resource):
    def get(self):
        return{}

class searchChannel(Resource):
    def get(self):
        return{}

#class for manipulating only one channel entry and its parameters#######################
class Records(Resource):
#gets data of all records
    def get(self):
        all_records = records.query.all()
        result = records_schema.dump(all_records)
        return jsonify(result)

#deletes chosen record from database
class delete_record(Resource):
    def delete(self, record):
        Record_info = records.query.get(record)
        db.session.delete(Record_info)
        db.session.commit()
        return record_schema.jsonify(Record_info)

##########################Resources and routes ###############################################################################################################################


api.add_resource(Frequency, "/frequency/<int:frequency>")
api.add_resource(allFrequencies, "/parameters")
api.add_resource(temperatureAndHumidity, "/tempAndHum")
api.add_resource(isThereSound, "/sound")
api.add_resource(isThereConnection, "/connection")
api.add_resource(searchChannel, "/search")
api.add_resource(Records, "/records")
api.add_resource(delete_record, "/records/<string:record>")

if __name__ == '__main__':
    app.run( host=config.host, debug= True, threaded=True, port=config.port)


    
