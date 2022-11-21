from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import csv
from datetime import datetime
import os
from pytz import timezone
import shutil


def copyFile(pathOriginal, pathTarget, fileName):
    original = pathOriginal + fileName + ".xlsx"
    target = pathTarget +  fileName + ".xlsx"
    shutil.copyfile(original, target)
    return 0


def createCSVfile(path, now , tunnelName):
    wb = Workbook()
    ws = wb.active
    fileName = tunnelName
    ws.title = now +  "-"+ tunnelName
    return wb, ws, fileName

def getTime():
    turkey = timezone('Europe/Istanbul')
    now = datetime.now(turkey)
    fileNow =now.strftime("%Y.%m.%d-%H.%M.%S")
    return fileNow

def deleteCSVfile(path , pathTarget):
    for file in os.listdir(path):  
        if (file.endswith('.xlsx')):
            try:
                os.remove(path+file)
            except Exception as e:
                print(e)
        else:
            pass

    for file in os.listdir(pathTarget):  
        if (file.endswith('.xlsx')):
            try:
                os.remove(pathTarget+file)
            except Exception as e:
                print(e)
        else:
            pass
    

class manageExcel():
    def __init__(self, wb,ws, fileName):
        self.wb = wb
        self.ws = ws
        self.fileName = fileName

    def appendHeader(self):
        wb = self.wb
        ws = self.ws
        fileName = self.fileName
        ws.append(['Date', 'Frequency', 'Validness', 'Signal Power', 'SNR', 'Temperature', 'Humidity'])
        wb.save(fileName +'.xlsx')

    def appendData(self, CSVdata):
        wb = self.wb
        ws = self.ws
        fileName = self.fileName
        for data in CSVdata:
            try:
                ws.append([data.date, data.freq, data.resp1, data.rssi, data.snr, data.temp, data.hum])
            except Exception as e:
                print(e)
        wb.save(fileName +'.xlsx')


def createDeleteCSV(path, pathTarget, tunnelName):
    deleteCSVfile(path, pathTarget)
    fileNow = getTime()
    wb, ws, fileName =createCSVfile(path, fileNow , tunnelName)
    return wb,ws, fileName

def addDataHeader(wb, ws, fileName, CSVdata):
    CSV = manageExcel(wb, ws, fileName)
    CSV.appendHeader()
    CSV.appendData(CSVdata)